from openpyxl.utils import get_column_letter

import myapp.post.cu_post as cu_post
import myapp.post.gs_post as gs_post

def make_excel_file(request):
    excel_file = request.FILES["excel_file"]
    store_type = request.POST["store_type"]

    # save post excel file
    excel_data = save_excel_file(excel_file, store_type)

    return excel_data
    


def save_excel_file(excel_file, store_type):
    # make post excel data
    gs_post_excel_data, cu_post_excel_data = get_excel_data(excel_file, store_type)

    # save post excel file
    gs_post.save_as_excel_file(gs_post_excel_data)
    cu_post.save_as_excel_file(cu_post_excel_data)

    return gs_post_excel_data

def get_excel_data(excel_file, store_type):
    import myapp.post.naver_post as naver_post
    import myapp.post.coupang_post as coupang_post
    import myapp.post.ably_post as ably_post

    if is_네이버(store_type):
        return naver_post.get_excel_data(excel_file)
    elif is_쿠팡(store_type):
        return coupang_post.get_excel_data(excel_file)
    elif is_에이블리(store_type):
        return ably_post.get_excel_data(excel_file)

    print('None')

    return None, None


def is_네이버(store_type):
    return store_type == '네이버'

def is_쿠팡(store_type):
    return store_type == '쿠팡'

def is_에이블리(store_type):
    return store_type == '에이블리'


def is_same_address(post_row1, post_row2):
    for i in range(0, 5):
        if post_row1[i] != post_row2[i]:
            return False
    return True


def get_same_address_index(post_excel_data, row_data):
    for i in range(0, len(post_excel_data)):
        if is_same_address(post_excel_data[i], row_data):
            return i
    return -1


def has_same_address(post_excel_data, row_data):
    if get_same_address_index(post_excel_data, row_data) != -1:
        return True
    return False


def get_index_by_row_and_column(sheet, row, column):
    return get_column_by_value(sheet, column) + row


def get_value_by_row_and_column(sheet, row, column):
    index = get_index_by_row_and_column(sheet, row, column)
    return get_value_by_index(sheet, index)


def get_value_by_index(sheet, index):
    if sheet[index].value != None:
        return str(sheet[index].value)
    else:
        return ' '


def get_column_by_value(sheet, value):
    for row in sheet.iter_rows():
        for cell in row:
            if (cell.value == value):
                return get_column_letter(cell.column)

    return None


