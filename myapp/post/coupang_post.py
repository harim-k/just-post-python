from openpyxl import load_workbook

from myapp.post.post import *


def get_excel_data(excel_file):
    # load excel file
    workbook = load_workbook(filename=excel_file)

    # getting a customer sheet
    sheet = workbook.worksheets[0]

    return make_excel_data(sheet)


def make_excel_data(sheet):
    # make post excel data
    gs_excel_data = list()
    cu_excel_data = list()

    # set first row from template
    gs_excel_data.append(gs_post.get_first_row_from_template())
    cu_excel_data.append(cu_post.get_first_row_from_template())

    excel_data = make_excel_value_data(sheet)
    gs_excel_data.extend(excel_data)
    cu_excel_data.extend(excel_data)

    return gs_excel_data, cu_excel_data


def make_excel_value_data(sheet):
    post_excel_data = list()

    # set value rows from store excel file
    for i in range(2, sheet.max_row + 1):
        row_data = list()
        index_string = str(i)

        수취인이름 = get_value_by_row_and_column(sheet, index_string, '수취인이름')
        우편번호 = get_value_by_row_and_column(sheet, index_string, '우편번호')
        수취인주소 = get_value_by_row_and_column(sheet, index_string, '수취인 주소')

        수취인전화번호 = get_value_by_row_and_column(sheet, index_string, '수취인전화번호')
        구매자전화번호 = get_value_by_row_and_column(sheet, index_string, '구매자전화번호')

        구매수 = get_value_by_row_and_column(sheet, index_string, '구매수(수량)')
        배송메세지 = get_value_by_row_and_column(sheet, index_string, '배송메세지')

        등록옵션명 = get_value_by_row_and_column(sheet, index_string, '등록옵션명')
        등록상품명 = get_value_by_row_and_column(sheet, index_string, '등록상품명')

        # 특수문자 제거
        등록옵션명 = 등록옵션명.replace('&', '')
        등록상품명 = 등록상품명.replace('&', '')

        # 단품 제거
        등록옵션명 = 등록옵션명.replace('단품', '')
        등록상품명 = 등록상품명.replace('단품', '')

        등록상품명 = 등록상품명.split(' ')[-1]

        품목 = 등록옵션명 if 등록옵션명 != '단일상품' else 등록상품명
        배송요청사항 = ' '.join([품목, 구매수, 배송메세지])
        지불방법 = '선불'

        row_data.append(수취인이름)  # 수취인명
        row_data.append(우편번호)  # 우편번호
        row_data.append(수취인주소)  # 주소 1
        row_data.append(수취인주소)  # 주소 2
        row_data.append(수취인전화번호)  # 전화번호 (수취인)
        row_data.append(구매자전화번호)  # 추가 전화번호
        row_data.append(배송요청사항)  # 배송요청사항
        row_data.append(지불방법)  # 지불방법

        # 같은 주소인 경우 하나로 합치기
        if has_same_address(post_excel_data, row_data):
            index = get_same_address_index(post_excel_data, row_data)
            # 배송요청사항에 품목, 수량 추가
            post_excel_data[index][6] = ' '.join([품목, 구매수, post_excel_data[index][6]])
        else:
            post_excel_data.append(row_data)

    return post_excel_data


def get_주소1(address):
    return address.split(get_주소2(address))[0]


def get_주소2(address):
    address = address.split('동 ')[-1]
    address = address.split('면 ')[-1]
    address = address.split('읍 ')[-1]
    return address
