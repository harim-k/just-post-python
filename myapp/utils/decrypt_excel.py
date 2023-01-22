from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from openpyxl import load_workbook

# The path to the encrypted Excel file
file_path = "encrypted.xlsx"

# The password used to encrypt the file
password = b"mypassword"

# Read in the encrypted file
with open(file_path, "rb") as f:
    data = f.read()

# Create the cipher object and set the key and IV
cipher = Cipher(algorithms.AES(password), modes.ECB(), backend=default_backend())
decryptor = cipher.decryptor()

# Decrypt the data
decrypted_data = decryptor.update(data) + decryptor.finalize()

# Write the decrypted data to a new file
with open("decrypted.xlsx", "wb") as f:
    f.write(decrypted_data)

# Load the decrypted Excel file
wb = load_workbook("decrypted.xlsx")

# Do something with the decrypted data
print(wb.get_sheet_names())





from openpyxl import load_workbook
from Cryptodome.Cipher import AES

# Define the encryption key and initialization vector
key = b"your_encryption_key"
iv = b"your_initialization_vector"

# Open the encrypted Excel file
with open("encrypted_file.xlsx", "rb") as f:
    # Read the contents of the file into memory
    data = f.read()

# Create a new AES cipher object
cipher = AES.new(key, AES.MODE_CBC, iv)

# Decrypt the data
decrypted_data = cipher.decrypt(data)

# Load the decrypted data into a new Excel workbook object
wb = load_workbook(filename=decrypted_data)

# Do something with the decrypted workbook
# ...
